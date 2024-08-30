import streamlit as st
from openpyxl import load_workbook
import io

def column_letter_to_index(letter):
    """
    Convert Excel column letter to 1-based index.
    E.g., 'A' -> 1, 'B' -> 2, ..., 'Z' -> 26, 'AA' -> 27
    """
    index = 0
    for char in letter:
        index = index * 26 + (ord(char.upper()) - ord('A') + 1)
    return index

def fun(file, row_number):
    wb = load_workbook(file)
    ws = wb.active

    # Convert column letter 'T' to index and insert new columns after it
    col_idx = column_letter_to_index('T') + 1  # Index for column U
    ws.insert_cols(col_idx, 4)  # Insert 4 columns

    # Rename new columns in row 10
    column_names = ["Closing Stock Qty", "EOS Qty", "MRP Sold Qty", "Anomaly"]
    for i, name in enumerate(column_names):
        ws.cell(row=10, column=col_idx + i, value=name)

    start_row = row_number + 1

    # Extract values from columns Z and AA
    z_aa_dict = {}
    for row in range(start_row, ws.max_row + 1):
        key = ws[f'Z{row}'].value
        value = ws[f'AA{row}'].value
        if key:
            z_aa_dict[key] = value

    # Update column U or add new rows in Q
    max_row = ws.max_row
    for row in range(start_row, max_row + 1):
        col_z_value = ws[f'Z{row}'].value
        col_q_values = [ws[f'Q{i}'].value for i in range(start_row, max_row + 1)]
        
        if col_z_value in col_q_values:
            col_q_index = col_q_values.index(col_z_value) + start_row
            ws[f'U{col_q_index}'] = z_aa_dict.get(col_z_value)
        else:
            max_row += 1
            ws[f'Q{max_row}'] = col_z_value
            ws[f'X{max_row}'] = 'X'
            ws[f'U{max_row}'] = z_aa_dict.get(col_z_value)

    # Set U value to 0 for Q values not in Z
    for row in range(start_row, max_row + 1):
        col_q_value = ws[f'Q{row}'].value
        if col_q_value not in z_aa_dict:
            ws[f'U{row}'] = 0

    # Process columns AD and AE
    ad_ae_dict = {}
    for row in range(start_row, max_row + 1):
        key = ws[f'AD{row}'].value
        value = ws[f'AE{row}'].value
        if key:
            ad_ae_dict[key] = value

    for row in range(start_row, max_row + 1):
        col_ad_value = ws[f'AD{row}'].value
        col_q_values = [ws[f'Q{i}'].value for i in range(start_row, max_row + 1)]
        
        if col_ad_value in col_q_values:
            col_q_index = col_q_values.index(col_ad_value) + start_row
            ws[f'V{col_q_index}'] = ad_ae_dict.get(col_ad_value)
        else:
            max_row += 1
            ws[f'Q{max_row}'] = col_ad_value
            ws[f'X{max_row}'] = 'Y'
            ws[f'V{max_row}'] = ad_ae_dict.get(col_ad_value)

    for row in range(start_row, max_row + 1):
        col_q_value = ws[f'Q{row}'].value
        if col_q_value not in ad_ae_dict:
            ws[f'V{row}'] = 0

    # Process columns AH and AI
    ah_ai_dict = {}
    for row in range(start_row, max_row + 1):
        key = ws[f'AH{row}'].value
        value = ws[f'AI{row}'].value
        if key:
            ah_ai_dict[key] = value

    for row in range(start_row, max_row + 1):
        col_ah_value = ws[f'AH{row}'].value
        col_q_values = [ws[f'Q{i}'].value for i in range(start_row, max_row + 1)]
        
        if col_ah_value in col_q_values:
            col_q_index = col_q_values.index(col_ah_value) + start_row
            ws[f'W{col_q_index}'] = ah_ai_dict.get(col_ah_value)
        else:
            max_row += 1
            ws[f'Q{max_row}'] = col_ah_value
            ws[f'X{max_row}'] = 'Z'
            ws[f'W{max_row}'] = ah_ai_dict.get(col_ah_value)

    for row in range(start_row, max_row + 1):
        col_q_value = ws[f'Q{row}'].value
        if col_q_value not in ah_ai_dict:
            ws[f'W{row}'] = 0

    # Save to a BytesIO object
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output

def main():
    uploaded_file = st.file_uploader("Upload your Excel file", type="xlsx")
    
    if uploaded_file:
        row_number = st.number_input("Enter the header row number:", min_value=1, value=10)
        
        if st.button("Process File"):
            # Process the file with the function
            processed_file = fun(uploaded_file, row_number)
            
            # Provide a download link for the processed file
            st.download_button(
                label="Download processed file",
                data=processed_file,
                file_name="processed_file.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            )

if __name__ == "__main__":
    main()
